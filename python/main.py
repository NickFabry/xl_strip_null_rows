#!/usr/bin/env python3
# -*- coding: utf-8 -*-



""" xl_strip_null_rows is designed to solve a specific problem - Excel sheets witha huge number of blank rows after the real data, often up to the maximum allowable number of rows, 1,048,576. Stripping out these empty rows in Excel is often difficult or impossible because of the massive size in memory such an Excel file takes, slowing manipulations to a crawl. This program aims to do that, stripping out blank rows in a much more efficient fashion, approximately 100x as fast as in Excel.

The algorthim used is limited. It iterates over all the Worksheets in the Workbook. In each Worksheet, it reads each cell in the first column sequentially from top to bottom. When it finds the FIRST blank/null cell, it then copies the all the rows above that blank/null cell, and puts them in a brand-new Workbook and identically named Worksheet. Note that all formatting,formulas, charts, etc. are not preserved in the copy; only the values. Note also that a stray blank row will cause the program to ignore all the rows(even if non-blank) below it. It also makes no attempt to strip away or detect empty columns. In the future, we may add additional options for these usecases.
"""


import sys, os
import argparse
import openpyxl



def arg_parser():
    parser = argparse.ArgumentParser(
        description='This program reads the provided Excel Workbook specified by --input, strips out all the rows after the first blank row, and writes the remaining rows to a new Excel Workbook. Note that everything (e.g. formatting, charts, etc.) other than the values are not preserved.')
    parser.add_argument('-i', '--input', metavar='FILE', 
        help='The Excel file you wish to strip blank rows from.')
    parser.add_argument('-o', '--output', metavar='FILE',
        help='The path you wish to write the output Excel file to.')
    parser.add_argument('-s', '--skip', metavar='NUMBER',
        type=int,
        help='The number of rows at the beginning you wish to skip blank testing. This is useful if a sheet contains a few blank rows near the top (e.g. for formatting purposes) that you want to preserve.')
    return parser



### MAIN ###
def main(args):
    iwb = openpyxl.load_workbook(args.input, read_only=True)
    owb = openpyxl.Workbook(write_only=True)
    for sn in iwb.sheetnames: 
        ows = owb.create_sheet(title=sn)
        i = 0
        for a_row in iwb[sn].iter_rows():
            # Test for empty row before empty cell to prevent Exception
            if (i > args.skip
            and (not bool(a_row) or a_row[0].value is None)):
                print(f'Sheet {sn} done at row {i}.')
                break
            else:
                row_vals = [c.value for c in a_row]
                ows.append(row_vals)
                i += 1
    print(f'Writing Workbook: {args.output}')
    owb.save(args.output)
    return locals()
### END MAIN ###



### MAIN RUNNER ###
if __name__ == '__main__':
    args = arg_parser().parse_args()
    print(args)
    if sys.__stdin__.isatty():
        locals().update(main(args))
    else:
        main(args)
### END ###